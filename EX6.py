import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os

class FormulaireStagiaire:
    def __init__(self, root):
        self.root = root
        self.root.title("Formulaire d'inscription - Stagiaire")
        self.root.geometry("500x550")
        self.root.resizable(False, False)

        title_label = tk.Label(root, text="FORMULAIRE D'INSCRIPTION", font=("Arial", 16, "bold"))
        title_label.pack(pady=10)

        main_frame = tk.Frame(root, padx=20, pady=10)
        main_frame.pack(fill="both", expand=True)

        fields = [
            ("Nom", 0),
            ("Prénom", 1),
            ("Date de naissance (JJ/MM/AAAA)", 2),
            ("N° Téléphone", 3),
            ("Email", 4),
            ("Adresse", 5),
            ("Niveau", 6)
        ]

        self.entries = {}

        for label_text, row in fields:
            label = tk.Label(main_frame, text=label_text + " :", font=("Arial", 10))
            label.grid(row=row, column=0, sticky="e", pady=8, padx=5)

            if label_text == "Niveau":
                self.niveau_var = tk.StringVar()
                niveau_options = ["Débutant", "Intermédiaire", "Avancé", "Expert"]
                entry = ttk.Combobox(main_frame, textvariable=self.niveau_var,
                                     values=niveau_options, width=35, state="readonly")
                entry.grid(row=row, column=1, pady=8, padx=5)
                self.entries[label_text] = self.niveau_var
            else:
                entry = tk.Entry(main_frame, width=38)
                entry.grid(row=row, column=1, pady=8, padx=5)
                self.entries[label_text] = entry

        submit_btn = tk.Button(main_frame, text="S'INSCRIRE", bg="#4CAF50", fg="white",
                               font=("Arial", 12, "bold"), command=self.enregistrer_donnees)
        submit_btn.grid(row=7, column=0, columnspan=2, pady=20)

        view_btn = tk.Button(main_frame, text="AFFICHER LES STAGIAIRES", bg="#2196F3", fg="white",
                             command=self.afficher_stagiaires)
        view_btn.grid(row=8, column=0, columnspan=2, pady=5)

        self.fichier_excel = "stagiaires.xlsx"
        self.creer_fichier_excel()

    def creer_fichier_excel(self):
        if not os.path.exists(self.fichier_excel):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Stagiaires"
            sheet.append(["Nom", "Prénom", "Date de naissance", "Téléphone",
                          "Email", "Adresse", "Niveau", "Date d'inscription"])
            workbook.save(self.fichier_excel)

    def enregistrer_donnees(self):
        nom = self.entries["Nom"].get().strip()
        prenom = self.entries["Prénom"].get().strip()
        date_naissance = self.entries["Date de naissance (JJ/MM/AAAA)"].get().strip()
        telephone = self.entries["N° Téléphone"].get().strip()
        email = self.entries["Email"].get().strip()
        adresse = self.entries["Adresse"].get().strip()
        niveau = self.entries["Niveau"].get().strip()

        # Vérification champs vides
        if not all([nom, prenom, date_naissance, telephone, email, adresse, niveau]):
            messagebox.showerror("Erreur", "Tous les champs doivent être remplis !")
            return

        # Validation date
        try:
            datetime.strptime(date_naissance, "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Erreur", "Format de date invalide (JJ/MM/AAAA)")
            return

        # Validation email
        if "@" not in email or "." not in email:
            messagebox.showerror("Erreur", "Email invalide !")
            return

        # Validation téléphone
        tel_clean = telephone.replace(" ", "").replace("-", "").replace("+", "")
        if not tel_clean.isdigit() or len(tel_clean) < 8:
            messagebox.showerror("Erreur", "Téléphone invalide !")
            return

        try:
            workbook = openpyxl.load_workbook(self.fichier_excel)
            sheet = workbook.active

            date_inscription = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            sheet.append([nom, prenom, date_naissance, telephone,
                          email, adresse, niveau, date_inscription])

            workbook.save(self.fichier_excel)

            messagebox.showinfo("Succès", f"{prenom} {nom} enregistré !")
            self.vider_formulaire()

        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    def vider_formulaire(self):
        for entry in self.entries.values():
            if isinstance(entry, tk.Entry):
                entry.delete(0, tk.END)
            else:
                entry.set("")

    def afficher_stagiaires(self):
        if not os.path.exists(self.fichier_excel):
            messagebox.showinfo("Info", "Aucun stagiaire.")
            return

        workbook = openpyxl.load_workbook(self.fichier_excel)
        sheet = workbook.active

        if sheet.max_row <= 1:
            messagebox.showinfo("Info", "Aucun stagiaire.")
            return

        view_window = tk.Toplevel(self.root)
        view_window.title("Liste des stagiaires")
        view_window.geometry("800x400")

        frame = tk.Frame(view_window)
        frame.pack(fill="both", expand=True)

        tree = ttk.Treeview(frame)
        tree.pack(fill="both", expand=True)

        columns = ["Nom", "Prénom", "Date naissance", "Téléphone", "Email", "Niveau", "Date inscription"]
        tree["columns"] = columns
        tree["show"] = "headings"

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120, anchor="center")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                tree.insert("", "end",
                            values=(row[0], row[1], row[2], row[3], row[4], row[6], row[7]))

if __name__ == "__main__":
    root = tk.Tk()
    app = FormulaireStagiaire(root)
    root.mainloop()